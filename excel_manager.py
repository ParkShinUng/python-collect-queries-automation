from config import Config
from typing import List, Tuple
from openpyxl import load_workbook
from datetime import datetime

        
class ExcelManager:
    def __init__(self, cfg: Config):
        self.cfg = cfg
        self.wb = load_workbook(self.cfg.excel_path)
        self.__ws = None
        self.sheet_list = self.wb.sheetnames
        
    @property
    def ws(self):
        return self.__ws
    
    @ws.setter
    def ws(self, value):
        self.__ws = self.wb[value]

    def read_jobs(self) -> List[Tuple[int, str]]:
        assert self.ws is not None
        jobs: List[Tuple[int, str]] = []
        max_row = self.ws.max_row

        for row in range(self.cfg.start_row, max_row + 1):
            cell_val = self.ws[f"{self.cfg.prompt_col}{row}"].value
            if cell_val:
                jobs.append((row, str(cell_val)))

        return jobs

    def apply_results(self, results: List[Tuple[int, str]]) -> None:
        """(row, formatted_queries) 리스트를 엑셀에 반영."""
        assert self.ws is not None
        
        def is_empty(value):
            return value is None or (isinstance(value, str) and value.strip() == "") or value == datetime.now().strftime("%Y/%m/%d")
        
        ws_result_col = self.cfg.result_check_start_col
        while not is_empty(self.ws.cell(row=self.cfg.start_col_row, column=ws_result_col).value):
            ws_result_col += 1
                
        for row, formatted in results:
            self.ws.cell(row=row, column=ws_result_col).value = formatted
            
    def save(self) -> None:
        assert self.wb is not None
        self.wb.save(self.cfg.excel_path)
    
    def set_worksheet(self, sheet_name: str) -> None:
        assert self.wb is not None
        self.ws = self.wb[sheet_name]
        
    def get_sheetnames(self) -> List:
        assert self.wb is not None
        return self.wb.sheetnames
    